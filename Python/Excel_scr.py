import bd_MDP_scr as bM
import pandas as pd
import os
import openpyxl as px
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill, Alignment, Side, Border
from datetime import datetime as dt 

def Full_rep(name_tabs_db, name_tabs):
    path_file = r'Отчетные файлы'
    path_shabl = r"C:\Users\bukre\OneDrive\Рабочий стол\ОДУ\Расчет МДП на ПР\Файлы Excel\shabl_MDP.xlsx"
    date = f"{dt.now().day}_{dt.now().month}_{dt.now().year}  {dt.now().hour}ч_{dt.now().minute}мин_{dt.now().second}сек"
    # Подключение к листу Excel
    file = px.load_workbook(path_shabl)
    fs = file.active
    # Перебор таблиц из БД
    for name_num in range(len(name_tabs_db)):
        max_row = fs.max_row
        
        # Занесение данных по всем схемам
        data_frame = pd.read_sql(f"SELECT * FROM {name_tabs_db[name_num]}", bM.bd)
        for r in dataframe_to_rows(data_frame, index=True, header=False):
            fs.append(r)
        max_row_new = fs.max_row
        # Форматирование данных
        # Заголовок схемы
        fs.merge_cells(f'B{max_row+1}:T{max_row+1}')
        name_scheme = fs[f'B{max_row+1}']
        name_scheme.value = f"{name_tabs[name_num]}"
        name_scheme.fill = PatternFill('solid', fgColor="00FFFF")
        fs[f'A{max_row+1}'].value = name_num + 1
        fs[f'A{max_row+1}'].fill = PatternFill('solid', fgColor="00FFFF")

        # Количество повторяемых строк с температурами
        rep_rows = data_frame['Temp'].value_counts()[0]
        
        # Объединение ячеек по температуре
        letters = ['A','B','C','D','E','F','G','H','I','J','K','S','T']
        m = 1
        for i in range(2+max_row,max_row_new+1,rep_rows):
            fs[f"A{i}"].value = f"{name_num + 1}.{m}"
            for let in letters:
                fs.merge_cells(f"{let}{i}:{let}{i + rep_rows - 1}")
            m += 1
            
        # Форматирование каждой ячейки
        for j in fs[f"A{max_row+1}:T{max_row_new}"]:
            for k in j:
                k.alignment = Alignment(horizontal="center", vertical="center")
                medium = Side(border_style="thin", color ="000000")
                k.border = Border(top=medium, bottom=medium, left=medium, right=medium)
        
    file.save(f"{path_file}\\МДП  {date}.xlsx")
    file.close()



# path_excel = r'Отчетные файлы'
# name_tab_db = "Norm_scheme"
# name = "Нормальная схема"
# num=1
# path_shabl = r"C:\Users\bukre\OneDrive\Рабочий стол\ОДУ\Расчет МДП на ПР\Файлы Excel\shabl_MDP.xlsx"
# Full_rep(name_tab_db, name, path_excel, path_shabl=path_shabl, number=num)
